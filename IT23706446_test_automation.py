from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_SHEET_NAME = " Test cases"
DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://www.pixelssuite.com/chat-translator")

DEFAULT_WAIT_MS = 15000
DEFAULT_RETRIES = 8
DEFAULT_RETRY_WAIT_MS = 1000
DEFAULT_TYPE_DELAY_MS = 200
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 500


def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass


def _resolve_path(p):
    if not p:
        return None
    path = Path(p)
    if path.is_absolute():
        return str(path)
    return str((ROOT_DIR / path).resolve())


def _normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _header_values(ws, row_index):
    return [ws.cell(row=row_index, column=c).value for c in range(1, ws.max_column + 1)]


def _find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        values = [_normalize_header(v) for v in _header_values(ws, r)]
        if "input" in values and "expectedoutput" in values:
            return r
    return 1


def _find_column_index(header_values, names):
    normalized = [_normalize_header(v) for v in header_values]
    for name in names:
        n = _normalize_header(name)
        if n in normalized:
            return normalized.index(n) + 1
    return None


def _merged_top_left_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell


def _set_cell_value(ws, row, col, value):
    _merged_top_left_cell(ws, row, col).value = value


def _clear_textarea(page, locator):
    locator.click()
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    locator.fill("")


def _read_output(page):
    """
    Try multiple strategies to read the translated output from the page.
    Strategy 1: second textarea (original assumption)
    Strategy 2: any visible div/p that looks like Sinhala output
    Strategy 3: JS evaluation scanning common output containers
    """
    # Strategy 1: second textarea
    try:
        textareas = page.locator("textarea").all()
        if len(textareas) >= 2:
            val = textareas[1].input_value()
            if val and val.strip():
                return val.strip()
    except Exception:
        pass

    # Strategy 2: JS scan for likely output containers
    try:
        result = page.evaluate("""() => {
            // Common output container selectors to try
            const selectors = [
                '#output', '.output', '[class*="output"]',
                '#result', '.result', '[class*="result"]',
                '#translation', '.translation', '[class*="translat"]',
                '[class*="sinhala"]', '[class*="target"]',
                'textarea:nth-of-type(2)',
                '.chat-output', '#chat-output',
                '[data-output]', '[data-result]'
            ];

            for (const sel of selectors) {
                const el = document.querySelector(sel);
                if (el) {
                    const text = (el.value || el.innerText || el.textContent || '').trim();
                    if (text.length > 0) return text;
                }
            }

            // Fallback: find the second large text container on the page
            const allTextareas = Array.from(document.querySelectorAll('textarea'));
            if (allTextareas.length >= 2) {
                return allTextareas[1].value.trim();
            }

            // Last resort: find any div/p with Sinhala-range Unicode characters
            const allEls = Array.from(document.querySelectorAll('div, p, span, pre'));
            for (const el of allEls) {
                const text = el.innerText?.trim() || '';
                // Sinhala Unicode range: \u0D80-\u0DFF
                if (/[\u0D80-\u0DFF]/.test(text) && text.length > 1) {
                    return text;
                }
            }

            return '';
        }""")
        if result and result.strip():
            return result.strip()
    except Exception:
        pass

    return ""


def _wait_for_output(page, previous_output="", timeout_ms=30000):
    """
    Wait for a non-empty, changed output to appear.
    Uses multiple detection strategies.
    """
    try:
        page.wait_for_function(
            """(previous) => {
                // Check if still loading/transliterating
                const buttons = Array.from(document.querySelectorAll('button'));
                const isLoading = buttons.some(btn =>
                    btn.innerText.toLowerCase().includes("transliterating") ||
                    btn.innerText.toLowerCase().includes("loading")
                );
                if (isLoading) return false;

                // Strategy 1: second textarea
                const textareas = Array.from(document.querySelectorAll('textarea'));
                if (textareas.length >= 2) {
                    const val = textareas[1].value?.trim() || '';
                    if (val.length > 0 && val !== previous) return true;
                }

                // Strategy 2: common output selectors
                const selectors = [
                    '#output', '.output', '[class*="output"]',
                    '#result', '.result', '[class*="result"]',
                    '#translation', '.translation', '[class*="translat"]',
                    '[class*="sinhala"]', '[class*="target"]',
                    '.chat-output', '#chat-output',
                    '[data-output]', '[data-result]'
                ];
                for (const sel of selectors) {
                    const el = document.querySelector(sel);
                    if (el) {
                        const text = (el.value || el.innerText || el.textContent || '').trim();
                        if (text.length > 0 && text !== previous) return true;
                    }
                }

                // Strategy 3: any Sinhala unicode text appeared
                const allEls = Array.from(document.querySelectorAll('div, p, span, pre, textarea'));
                for (const el of allEls) {
                    const text = (el.value || el.innerText || '').trim();
                    if (/[\u0D80-\u0DFF]/.test(text) && text !== previous && text.length > 1) {
                        return true;
                    }
                }

                return false;
            }""",
            arg=previous_output,
            timeout=timeout_ms
        )
        return True
    except Exception:
        return False


def _debug_page_structure(page):
    """Print the page's textarea/output structure to help diagnose issues."""
    try:
        info = page.evaluate("""() => {
            const textareas = Array.from(document.querySelectorAll('textarea'));
            const info = {
                textarea_count: textareas.length,
                textarea_values: textareas.map((t, i) => ({
                    index: i,
                    value_length: t.value.length,
                    value_preview: t.value.substring(0, 50),
                    id: t.id,
                    class: t.className,
                    placeholder: t.placeholder
                })),
                buttons: Array.from(document.querySelectorAll('button')).map(b => b.innerText.trim()).filter(t => t)
            };
            return info;
        }""")
        print(f"  [DEBUG] Page structure: {info}")
    except Exception as e:
        print(f"  [DEBUG] Could not inspect page: {e}")


def _parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS)
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS)
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS)
    parser.add_argument("--save-every", type=int, default=1)
    parser.add_argument("--keep-open", action="store_true")
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--debug", action="store_true", help="Print page structure info for first row")
    return parser.parse_args()


def run_test():
    _configure_stdout()
    args = _parse_args()

    excel_path = _resolve_path(args.excel)

    if not os.path.exists(excel_path):
        print(f"Error: File '{excel_path}' not found.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active

    header_row = _find_header_row(ws)
    header_values = _header_values(ws, header_row)

    input_col = _find_column_index(header_values, ["Input"])
    expected_col = _find_column_index(header_values, ["Expected output", "Expected Output"])
    actual_col = _find_column_index(header_values, ["Actual output", "Actual Output"])
    status_col = _find_column_index(header_values, ["Status"])

    if not input_col:
        print("Error: Input column not found.")
        return

    if not expected_col:
        print("Error: Expected output column not found.")
        return

    if not actual_col:
        actual_col = ws.max_column + 1
        ws.cell(row=header_row, column=actual_col).value = "Actual output"

    if not status_col:
        status_col = ws.max_column + 1
        ws.cell(row=header_row, column=status_col).value = "Status"

    rows_total = ws.max_row - header_row
    print(f"Starting Frontend-Only test with {rows_total} rows...")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=args.headless,
            slow_mo=max(0, int(args.slow_mo_ms))
        )

        page = browser.new_page()
        page.set_default_timeout(60000)

        try:
            page.goto(args.url, wait_until="domcontentloaded")
            page.wait_for_selector("textarea", timeout=60000)
            print("Frontend loaded successfully.")
        except Exception as e:
            print(f"Error loading frontend: {e}")
            browser.close()
            return

        # Wait a bit extra for the page to fully render JS components
        page.wait_for_timeout(2000)

        input_locator = page.locator("textarea").nth(0)
        action_locator = page.get_by_role("button", name=re.compile("Transliterate", re.I))

        processed = 0
        first_row = True
        last_input = None

        for row_index in range(header_row + 1, ws.max_row + 1):
            input_value = _merged_top_left_cell(ws, row_index, input_col).value
            if not input_value:
                continue

            singlish_input = str(input_value).strip()

            if singlish_input == last_input:
                continue

            last_input = singlish_input

            expected_value = _merged_top_left_cell(ws, row_index, expected_col).value
            expected_output = str(expected_value).strip() if expected_value else ""

            print(f"Testing [Row {row_index}]: {singlish_input}")

            try:
                # Read current output before typing (to detect change)
                previous_output = _read_output(page)

                # Clear output box before each test
                try:
                    page.locator("textarea").nth(1).evaluate("el => el.value = ''")
                except Exception:
                    pass

                _clear_textarea(page, input_locator)
                page.wait_for_timeout(300)

                input_locator.type(
                    singlish_input,
                    delay=max(0, int(args.type_delay_ms))
                )

                # Small pause after typing before clicking
                page.wait_for_timeout(500)

                action_locator.click()

                # Debug page structure on first row or if --debug flag set
                if first_row or args.debug:
                    page.wait_for_timeout(1000)
                    _debug_page_structure(page)
                    first_row = False

                output_ready = _wait_for_output(page, previous_output, timeout_ms=30000)

                if not output_ready:
                    print("  -> Warning: Output did not appear within 30 seconds")

                # Extra wait for slow responses
                page.wait_for_timeout(max(0, int(args.wait_ms)))

                actual_output = _read_output(page)

                _set_cell_value(ws, row_index, actual_col, actual_output)

                if expected_output:
                    status = "PASS" if actual_output == expected_output else "FAIL"
                else:
                    status = "COLLECTED"

                _set_cell_value(ws, row_index, status_col, status)

                print(f"  -> Actual output: {actual_output}")
                print(f"  -> {status}")

                processed += 1

                if args.save_every and processed % args.save_every == 0:
                    wb.save(excel_path)

            except Exception as e:
                print(f"Error in UI interaction: {e}")
                _set_cell_value(ws, row_index, status_col, "UI Error")
                wb.save(excel_path)

        wb.save(excel_path)

        if args.keep_open and not args.headless:
            print("Keeping browser open. Press CTRL+C to stop.")
            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                wb.save(excel_path)

        browser.close()

    print(f"Test completed. Results saved to {excel_path}")


if __name__ == "__main__":
    run_test()